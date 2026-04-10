#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
外汇衍生品交易字段导出脚本
生成结构化的Excel文件
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel():
    # 创建Excel writer
    output_path = "wiki/work/外汇衍生品交易字段汇总.xlsx"
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # ========== Sheet 1: 基础交易字段（所有品种共有） ==========
        base_fields = [
            ["字段类别", "字段名称", "英文名称", "数据类型", "是否必填", "说明/示例"],
            ["交易基本信息", "交易日期", "Trade Date", "Date", "是", "交易达成日期，如：2026-04-08"],
            ["交易基本信息", "起息日期", "Value Date", "Date", "是", "资金交割日期，Spot为T+2"],
            ["交易基本信息", "到期日期", "Maturity Date", "Date", "条件", "合约到期日期，适用于远期/掉期/期权"],
            ["交易基本信息", "交易对手", "Counterparty", "String", "是", "交易对手方名称，如：Bank A"],
            ["交易基本信息", "账簿/组合", "Book/Portfolio", "String", "是", "内部簿记归属"],
            ["交易基本信息", "交易编号", "Trade ID", "String", "系统生成", "系统生成的唯一标识，如：FX_20260408_001"],
            ["交易基本信息", "交易状态", "Trade Status", "Enum", "系统维护", "NEW/VALIDATED/CONFIRMED/LIVE/SETTLED"],
            
            ["货币信息", "买入货币", "Buy Currency", "String(3)", "是", "买入货币代码，如：CNY"],
            ["货币信息", "卖出货币", "Sell Currency", "String(3)", "是", "卖出货币代码，如：USD"],
            ["货币信息", "买入金额", "Buy Amount", "Decimal", "是", "买入货币数量，如：71,500,000"],
            ["货币信息", "卖出金额", "Sell Amount", "Decimal", "是", "卖出货币数量，如：10,000,000"],
            ["货币信息", "汇率", "Exchange Rate", "Decimal", "是", "交易约定汇率，如：7.1500"],
            ["货币信息", "货币对", "Currency Pair", "String(6)", "系统生成", "如：USDCNY"],
            
            ["结算信息", "结算指令", "Settlement Instructions", "Text", "是", "资金交割账户信息(SDI)"],
            ["结算信息", "结算方式", "Settlement Method", "Enum", "是", "CLS/双边结算/净额结算"],
            ["结算信息", "参考源", "Reference Sources", "String", "否", "报价来源，如：Bloomberg/Reuters"],
            ["结算信息", "特殊条款", "Special Conditions", "Text", "否", "特殊约定条款"],
            
            ["风控信息", "交易对手限额", "Counterparty Limit", "Decimal", "系统检查", "剩余额度：50M USD"],
            ["风控信息", "货币对限额", "Currency Pair Limit", "Decimal", "系统检查", "当前敞口：80M USD"],
            ["风控信息", "NOP限额", "NOP Limit", "Decimal", "系统检查", "净隔夜头寸限额：120M USD"],
        ]
        
        df_base = pd.DataFrame(base_fields[1:], columns=base_fields[0])
        df_base.to_excel(writer, sheet_name='1-基础交易字段', index=False)
        
        # ========== Sheet 2: 外汇即期（FX Spot） ==========
        spot_fields = [
            ["字段名称", "英文名称", "数据类型", "示例值", "说明"],
            ["交易日期", "Trade Date", "Date", "2026-04-08", "T日"],
            ["起息日期", "Value Date", "Date", "2026-04-10", "T+2交割"],
            ["买入货币", "Buy Currency", "String", "CNY", ""],
            ["卖出货币", "Sell Currency", "String", "USD", ""],
            ["买入金额", "Buy Amount", "Decimal", "71,500,000", "CNY金额"],
            ["卖出金额", "Sell Amount", "Decimal", "10,000,000", "USD金额"],
            ["汇率", "Exchange Rate", "Decimal", "7.1500", "即期汇率"],
            ["交易对手", "Counterparty", "String", "Bank A", ""],
            ["确认报文", "Confirmation Msg", "String", "SWIFT MT300", "外汇确认标准报文"],
        ]
        df_spot = pd.DataFrame(spot_fields[1:], columns=spot_fields[0])
        df_spot.to_excel(writer, sheet_name='2-外汇即期', index=False)
        
        # ========== Sheet 3: 外汇远期（FX Forward） ==========
        forward_fields = [
            ["字段名称", "英文名称", "数据类型", "示例值", "说明"],
            ["交易日期", "Trade Date", "Date", "2026-04-08", "T日"],
            ["到期日期", "Maturity Date", "Date", "2026-07-10", "3个月远期"],
            ["买入", "Buy", "String", "USD 5,000,000", "买入货币及金额"],
            ["卖出", "Sell", "String", "CNY 36,100,000", "卖出货币及金额"],
            ["远期汇率", "Forward Rate", "Decimal", "7.2200", "约定远期汇率"],
            ["市场远期汇率", "Market Forward Rate", "Decimal", "7.2250", "用于盯市"],
            ["估值损益", "MTM P&L", "Decimal", "+25,000", "盯市浮动盈亏(CNY)"],
            ["累计损益", "Accumulated P&L", "Decimal", "+150,000", "累计盯市盈亏"],
        ]
        df_forward = pd.DataFrame(forward_fields[1:], columns=forward_fields[0])
        df_forward.to_excel(writer, sheet_name='3-外汇远期', index=False)
        
        # ========== Sheet 4: 外汇掉期（FX Swap） ==========
        swap_fields = [
            ["字段名称", "英文名称", "数据类型", "近端示例", "远端示例", "说明"],
            ["起息日", "Value Date", "Date", "2026-04-10", "2026-07-10", ""],
            ["买入", "Buy", "String", "USD 10M", "CNY 72.2M", "近端买入USD，远端买入CNY"],
            ["卖出", "Sell", "String", "CNY 71.5M", "USD 10M", "近端卖出CNY，远端卖出USD"],
            ["汇率", "Exchange Rate", "Decimal", "7.1500", "7.2200", "近端即期汇率，远端远期汇率"],
            ["掉期点收入", "Swap Point Income", "Decimal", "700,000 CNY", "", "72.2M - 71.5M = 0.7M"],
        ]
        df_swap = pd.DataFrame(swap_fields[1:], columns=swap_fields[0])
        df_swap.to_excel(writer, sheet_name='4-外汇掉期', index=False)
        
        # ========== Sheet 5: 无本金交割远期（NDF） ==========
        ndf_fields = [
            ["字段名称", "英文名称", "数据类型", "示例值", "说明"],
            ["合约汇率", "Contract Rate", "Decimal", "7.2200", "约定的NDF汇率"],
            ["参考汇率", "Reference Rate", "Decimal", "7.2300", "到期日确定的参考汇率"],
            ["名义本金", "Notional Amount", "Decimal", "10,000,000", "USD金额"],
            ["差额", "Difference", "Decimal", "100,000 CNY", "(7.2300-7.2200)×10M"],
            ["结算方向", "Settlement Direction", "Enum", "我方支付", "根据差额正负确定"],
        ]
        df_ndf = pd.DataFrame(ndf_fields[1:], columns=ndf_fields[0])
        df_ndf.to_excel(writer, sheet_name='5-无本金交割远期', index=False)
        
        # ========== Sheet 6: 货币掉期（Currency Swap） ==========
        ccs_fields = [
            ["字段名称", "英文名称", "数据类型", "USD端", "CNY端", "说明"],
            ["名义本金", "Notional Amount", "Decimal", "50M USD", "357.5M CNY", "按即期汇率换算"],
            ["利率类型", "Rate Type", "Enum", "固定4.5%", "固定3.0%", "也可以是浮动利率"],
            ["付息频率", "Payment Frequency", "Enum", "半年", "半年", "利息交换周期"],
            ["期初交换", "Initial Exchange", "String", "收USD 50M", "付CNY 357.5M", "期初本金交换"],
            ["期间利息", "Periodic Interest", "String", "付1.125M", "收5.3625M", "半年利息"],
            ["期末交换", "Final Exchange", "String", "付USD 50M+利息", "收CNY 357.5M+利息", "期末本金+利息返还"],
        ]
        df_ccs = pd.DataFrame(ccs_fields[1:], columns=ccs_fields[0])
        df_ccs.to_excel(writer, sheet_name='6-货币掉期', index=False)
        
        # ========== Sheet 7: 外汇期权（FX Option） ==========
        option_fields = [
            ["字段类别", "字段名称", "英文名称", "数据类型", "示例值", "说明"],
            ["基础信息", "期权类型", "Option Type", "Enum", "欧式看涨USD Call/CNY Put", "Call/Put"],
            ["基础信息", "名义本金", "Notional Amount", "Decimal", "10M USD", "期权面值"],
            ["基础信息", "期权费", "Premium", "Decimal", "350,000 CNY", "买方支付给卖方的费用"],
            ["基础信息", "币种", "Currency", "String", "CNY", "期权费币种"],
            
            ["行权条款", "行权价", "Strike Price", "Decimal", "7.2500", "约定行权汇率"],
            ["行权条款", "到期日", "Expiry Date", "Date", "2026-07-08", "期权到期日"],
            ["行权条款", "行权方式", "Exercise Style", "Enum", "欧式", "欧式/美式/百慕大"],
            ["行权条款", "交割方式", "Settlement Type", "Enum", "现金交割", "现金/实物交割"],
            
            ["希腊字母", "Delta", "Delta", "Decimal", "0.43 → 0.85", "标的价格变动1单位的期权价值变化"],
            ["希腊字母", "Gamma", "Gamma", "Decimal", "0.15 → 0.08", "Delta对价格的敏感度"],
            ["希腊字母", "Vega", "Vega", "Decimal", "-", "波动率变动1%的期权价值变化"],
            ["希腊字母", "Theta", "Theta", "Decimal", "-", "时间衰减"],
            ["希腊字母", "Rho", "Rho", "Decimal", "-", "利率变动1%的期权价值变化"],
            ["希腊字母", "期权市值", "Option Market Value", "Decimal", "368,000 → 1,250,000", "随标的价格变化"],
            
            ["行权结果", "内在价值", "Intrinsic Value", "Decimal", "300,000 CNY", "(7.2800-7.2500)×10M"],
            ["行权结果", "到期汇率", "Spot at Expiry", "Decimal", "7.2800", "到期日即期汇率"],
            ["行权结果", "是否行权", "Exercise Flag", "Boolean", "是", "价内期权自动行权"],
        ]
        df_option = pd.DataFrame(option_fields[1:], columns=option_fields[0])
        df_option.to_excel(writer, sheet_name='7-外汇期权', index=False)
        
        # ========== Sheet 8: 外币拆借（FX Money Market） ==========
        mm_fields = [
            ["字段名称", "英文名称", "数据类型", "示例值", "说明"],
            ["借入金额", "Borrow Amount", "Decimal", "20M USD", "拆借本金"],
            ["利率基准", "Rate Index", "String", "SOFR", "基准利率"],
            ["利差", "Spread", "Decimal", "50bps", "0.50%"],
            ["总利率", "All-in Rate", "Decimal", "SOFR+0.5%", "实际执行利率"],
            ["期限", "Tenor", "String", "1个月", ""],
            ["起息日", "Value Date", "Date", "2026-04-10", ""],
            ["到期日", "Maturity Date", "Date", "2026-05-10", ""],
            ["每日应计利息", "Daily Accrued Interest", "Formula", "本金×(SOFR+0.5%)/360", ""],
            ["到期利息", "Total Interest", "Decimal", "81,166.67 USD", "1个月利息总额"],
            ["到期本息", "Principal + Interest", "Decimal", "20,081,166.67 USD", "本金+利息"],
        ]
        df_mm = pd.DataFrame(mm_fields[1:], columns=mm_fields[0])
        df_mm.to_excel(writer, sheet_name='8-外币拆借', index=False)
        
        # ========== Sheet 9: 字段汇总对照表 ==========
        summary_fields = [
            ["品种", "特有字段数量", "关键特有字段", "生命周期特点"],
            ["外汇即期", "0", "无", "T+2交割，一次性生命周期"],
            ["外汇远期", "2", "远期汇率、每日盯市损益", "每日盯市，到期结算"],
            ["外汇掉期", "4", "近端/远端汇率、掉期点收入", "近端+远端两笔交易"],
            ["无本金交割远期", "3", "参考汇率、差额、名义本金", "差额结算，无本金交割"],
            ["货币掉期", "5", "两端利率、付息频率、本金交换", "期初期末本金交换，定期利息交换"],
            ["外汇期权", "10+", "行权价、期权费、希腊字母", "每日盯市，到期行权/放弃"],
            ["外币拆借", "4", "利率基准、利差、应计利息", "每日计息，到期还本付息"],
        ]
        df_summary = pd.DataFrame(summary_fields[1:], columns=summary_fields[0])
        df_summary.to_excel(writer, sheet_name='9-字段汇总对照', index=False)
        
        # 获取workbook进行格式设置
        workbook = writer.book
        
        # 设置所有sheet的格式
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 应用表头样式
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 设置列宽和数据格式
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # 自动调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 冻结首行
            worksheet.freeze_panes = "A2"
    
    print(f"Excel文件已生成: {output_path}")
    return output_path

if __name__ == "__main__":
    create_excel()
